import os.path

import openpyxl


def create_workbook(name):
    wb = openpyxl.Workbook
    wb.active.title = 'Sheet1'        # renames the active sheet
    # wb.create_sheet('Sheet2')       # create a 2nd sheet
    wb.save(name)

    return wb


def read_sheet(sheet):
    rows = sheet.iter_rows()       # return generator class
    # print(type(rows))

    for row in rows:
        # print(row)
        for cell in row:
            print(cell.value)


def find_and_replace(wb, old, new):

    for sheet_name in wb.sheetnames:

        rows = wb[sheet_name].iter_rows()       # return generator class

        for row in rows:
            # print(row)
            for cell in row:
                if cell.value == old:
                    cell.value = new

    return wb


if __name__ == '__main__':
    # wb = create_workbook('test_spreadsheet.xlsx')
    file = 'test_spreadsheet.xlsx'

    # check if file exists
    if os.path.isfile(file):
        # load workbook
        wb = openpyxl.load_workbook(file)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        print(ws.dimensions)

        # replace null values in all sheets
        # wb = find_and_replace(wb, '[null]', '')

        # save workbook
        # wb.save(file)
    else:
        print(f"{file} is not a valid file. Please try again.")
